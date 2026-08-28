import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

wb = openpyxl.Workbook()
# remove default sheet
wb.remove(wb.active)

# Fonts & Colors
font_family = "Arial"
header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
title_font = Font(name=font_family, size=14, bold=True, color="0F766E")
kpi_title_font = Font(name=font_family, size=10, bold=True, color="475569")
kpi_val_font = Font(name=font_family, size=13, bold=True, color="0F766E")
bold_font = Font(name=font_family, size=11, bold=True)
regular_font = Font(name=font_family, size=10)
italic_font = Font(name=font_family, size=9, italic=True, color="64748B")

# Fills
teal_dark_fill = PatternFill(start_color="0F766E", end_color="0F766E", fill_type="solid")
teal_light_fill = PatternFill(start_color="CCFBF1", end_color="CCFBF1", fill_type="solid")
slate_header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
gray_header_fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
highlight_yellow = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
green_soft_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")

# Borders
thin_side = Side(style='thin', color='CBD5E1')
med_side = Side(style='medium', color='94A3B8')
double_side = Side(style='double', color='0F766E')
all_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
total_border = Border(top=thin_side, bottom=double_side, left=thin_side, right=thin_side)

# Alignments
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
right_align = Alignment(horizontal='right', vertical='center')
left_align = Alignment(horizontal='left', vertical='center')

# ==============================================================================
# SHEET 1: المندوبين والمحافظات
# ==============================================================================
ws1 = wb.create_sheet(title="المندوبين والمحافظات")
ws1.views.sheetView[0].rightToLeft = True

ws1.append(["المحافظة", "المندوب", "كود المندوب", "ملاحظات"])
for col_num in range(1, 5):
    cell = ws1.cell(1, col_num)
    cell.font = header_font
    cell.fill = teal_dark_fill
    cell.alignment = center_align
    cell.border = all_border
ws1.row_dimensions[1].height = 28

delegates_data = [
    # الكوت
    ("الكوت", "مذخر الكوت مباشر", "KUT-DIR", "توزيع مباشر"),
    ("الكوت", "1-Kut", "KUT-01", "مندوب منطقة الكوت 1"),
    ("الكوت", "2-Kut", "KUT-02", "مندوب منطقة الكوت 2"),
    ("الكوت", "3-Kut", "KUT-03", "مندوب منطقة الكوت 3"),
    ("الكوت", "4-Kut", "KUT-04", "مندوب منطقة الكوت 4"),
    ("الكوت", "5-Kut", "KUT-05", "مندوب منطقة الكوت 5"),
    ("الكوت", "6-Kut", "KUT-06", "مندوب منطقة الكوت 6"),
    ("الكوت", "7-Kut", "KUT-07", "مندوب منطقة الكوت 7"),
    # العمارة
    ("العمارة", "مذخر العمارة مباشر", "AMR-DIR", "توزيع مباشر"),
    ("العمارة", "1-Amara", "AMR-01", "مندوب منطقة العمارة 1"),
    ("العمارة", "2-Amara", "AMR-02", "مندوب منطقة العمارة 2"),
    ("العمارة", "3-Amara", "AMR-03", "مندوب منطقة العمارة 3"),
    ("العمارة", "4-Amara", "AMR-04", "مندوب منطقة العمارة 4"),
    ("العمارة", "5-Amara", "AMR-05", "مندوب منطقة العمارة 5"),
    ("العمارة", "6-Amara", "AMR-06", "مندوب منطقة العمارة 6"),
    ("العمارة", "7-Amara", "AMR-07", "مندوب منطقة العمارة 7"),
    # البصرة
    ("البصرة", "مذخر البصرة مباشر", "BAS-DIR", "توزيع مباشر"),
    ("البصرة", "1-Basra", "BAS-01", "مندوب منطقة البصرة 1"),
    ("البصرة", "2-Basra", "BAS-02", "مندوب منطقة البصرة 2"),
    ("البصرة", "3-Basra", "BAS-03", "مندوب منطقة البصرة 3"),
    ("البصرة", "4-Basra", "BAS-04", "مندوب منطقة البصرة 4"),
    ("البصرة", "5-Basra", "BAS-05", "مندوب منطقة البصرة 5"),
    ("البصرة", "6-Basra", "BAS-06", "مندوب منطقة البصرة 6"),
    ("البصرة", "7-Basra", "BAS-07", "مندوب منطقة البصرة 7"),
    # الناصرية
    ("الناصرية", "مذخر الناصرية مباشر", "NAS-DIR", "توزيع مباشر"),
    ("الناصرية", "1-Nasiriya", "NAS-01", "مندوب منطقة الناصرية 1"),
    ("الناصرية", "2-Nasiriya", "NAS-02", "مندوب منطقة الناصرية 2"),
    ("الناصرية", "3-Nasiriya", "NAS-03", "مندوب منطقة الناصرية 3"),
    ("الناصرية", "4-Nasiriya", "NAS-04", "مندوب منطقة الناصرية 4"),
    ("الناصرية", "5-Nasiriya", "NAS-05", "مندوب منطقة الناصرية 5"),
    ("الناصرية", "6-Nasiriya", "NAS-06", "مندوب منطقة الناصرية 6"),
    ("الناصرية", "7-Nasiriya", "NAS-07", "مندوب منطقة الناصرية 7"),
]

for row_idx, r in enumerate(delegates_data, start=2):
    ws1.append(list(r))
    fill_to_use = zebra_fill if row_idx % 2 == 0 else PatternFill(fill_type=None)
    for col_num in range(1, 5):
        cell = ws1.cell(row_idx, col_num)
        cell.font = regular_font
        cell.alignment = center_align if col_num in [1, 3] else right_align
        cell.border = all_border
        if fill_to_use.fill_type:
            cell.fill = fill_to_use
    ws1.row_dimensions[row_idx].height = 22

# ==============================================================================
# SHEET 2: المواد والأسعار
# ==============================================================================
ws2 = wb.create_sheet(title="المواد والأسعار")
ws2.views.sheetView[0].rightToLeft = True

ws2.append(["المادة", "السعر المفرد (د.ع)", "الشركة الموردة", "ملاحظات"])
for col_num in range(1, 5):
    cell = ws2.cell(1, col_num)
    cell.font = header_font
    cell.fill = teal_dark_fill
    cell.alignment = center_align
    cell.border = all_border
ws2.row_dimensions[1].height = 28

items_data = [
    ("Paracetamol", 1200, "LDP", "أقراص / شراب"),
    ("Ibuprofen", 1500, "LDP", "أقراص 400 ملغم"),
    ("Amoxicillin", 2500, "MEDREICH", "كبسول 500 ملغم"),
    ("Azithromycin", 3000, "LDP", "أقراص 500 ملغم"),
    ("Omeprazole", 2000, "MEDREICH", "كبسول 20 ملغم"),
    ("Metformin", 1500, "MEDREICH", "أقراص 500 ملغم"),
    ("Amlodipine", 2500, "LDP", "أقراص 5 ملغم"),
    ("Atorvastatin", 3500, "MEDREICH", "أقراص 20 ملغم"),
    ("Losartan", 3000, "LDP", "أقراص 50 ملغم"),
    ("Levothyroxine", 4000, "MEDREICH", "أقراص 50 ميكروغرام"),
    ("Salbutamol", 2000, "LDP", "بخاخ / شراب"),
    ("Pantoprazole", 2500, "MEDREICH", "أقراص 40 ملغم"),
    ("Ciprofloxacin", 1500, "LDP", "أقراص 500 ملغم"),
    ("Diclofenac", 1000, "LDP", "أمبول / أقراص 50 ملغم"),
    ("Cetirizine", 1500, "MEDREICH", "أقراص 10 ملغم"),
    ("Loratadine", 2000, "MEDREICH", "أقراص 10 ملغم"),
]

for row_idx, r in enumerate(items_data, start=2):
    ws2.append(list(r))
    fill_to_use = zebra_fill if row_idx % 2 == 0 else PatternFill(fill_type=None)
    for col_num in range(1, 5):
        cell = ws2.cell(row_idx, col_num)
        cell.font = regular_font
        cell.alignment = center_align if col_num in [3, 4] else right_align
        cell.border = all_border
        if fill_to_use.fill_type:
            cell.fill = fill_to_use
        if col_num == 2:
            cell.number_format = '#,##0 "د.ع"'
            cell.font = bold_font
    ws2.row_dimensions[row_idx].height = 22

# ==============================================================================
# SHEET 3: تارجت المحافظات
# ==============================================================================
ws3 = wb.create_sheet(title="تارجت المحافظات")
ws3.views.sheetView[0].rightToLeft = True

ws3.append(["السنة", "الشهر", "المحافظة", "الهدف / التارغت (د.ع)", "ملاحظات الخطة"])
for col_num in range(1, 6):
    cell = ws3.cell(1, col_num)
    cell.font = header_font
    cell.fill = teal_dark_fill
    cell.alignment = center_align
    cell.border = all_border
ws3.row_dimensions[1].height = 28

target_data = [
    # 2026 Monthly Targets
    (2026, 8, "الكوت", 25000000, "هدف شهر آب 2026"),
    (2026, 8, "العمارة", 20000000, "هدف شهر آب 2026"),
    (2026, 8, "البصرة", 35000000, "هدف شهر آب 2026"),
    (2026, 8, "الناصرية", 22000000, "هدف شهر آب 2026"),
    
    (2026, 9, "الكوت", 28000000, "هدف شهر أيلول 2026"),
    (2026, 9, "العمارة", 22000000, "هدف شهر أيلول 2026"),
    (2026, 9, "البصرة", 40000000, "هدف شهر أيلول 2026"),
    (2026, 9, "الناصرية", 25000000, "هدف شهر أيلول 2026"),

    (2026, 10, "الكوت", 30000000, "هدف شهر تشرين الأول 2026"),
    (2026, 10, "العمارة", 25000000, "هدف شهر تشرين الأول 2026"),
    (2026, 10, "البصرة", 45000000, "هدف شهر تشرين الأول 2026"),
    (2026, 10, "الناصرية", 28000000, "هدف شهر تشرين الأول 2026"),
]

for row_idx, r in enumerate(target_data, start=2):
    ws3.append(list(r))
    fill_to_use = zebra_fill if row_idx % 2 == 0 else PatternFill(fill_type=None)
    for col_num in range(1, 6):
        cell = ws3.cell(row_idx, col_num)
        cell.font = regular_font
        cell.alignment = center_align if col_num in [1, 2, 3] else right_align
        cell.border = all_border
        if fill_to_use.fill_type:
            cell.fill = fill_to_use
        if col_num == 4:
            cell.number_format = '#,##0 "د.ع"'
            cell.font = bold_font
    ws3.row_dimensions[row_idx].height = 22

# Add Governorates Data Validation for Sheet 3
gov_validation = DataValidation(type="list", formula1='"الكوت,العمارة,البصرة,الناصرية"', allow_blank=True)
ws3.add_data_validation(gov_validation)
gov_validation.add(f"C2:C100")

# ==============================================================================
# SHEETS 4-7: صفحات سجلات المحافظات (الكوت, العمارة, البصرة, الناصرية)
# ==============================================================================
gov_configs = [
    ("سجلات الكوت", "الكوت", "مذخر الكوت مباشر,1-Kut,2-Kut,3-Kut,4-Kut,5-Kut,6-Kut,7-Kut", 25000000),
    ("سجلات العمارة", "العمارة", "مذخر العمارة مباشر,1-Amara,2-Amara,3-Amara,4-Amara,5-Amara,6-Amara,7-Amara", 20000000),
    ("سجلات البصرة", "البصرة", "مذخر البصرة مباشر,1-Basra,2-Basra,3-Basra,4-Basra,5-Basra,6-Basra,7-Basra", 35000000),
    ("سجلات الناصرية", "الناصرية", "مذخر الناصرية مباشر,1-Nasiriya,2-Nasiriya,3-Nasiriya,4-Nasiriya,5-Nasiriya,6-Nasiriya,7-Nasiriya", 22000000),
]

company_val = DataValidation(type="list", formula1='"LDP,MEDREICH"', allow_blank=True)
item_val = DataValidation(type="list", formula1="'المواد والأسعار'!$A$2:$A$17", allow_blank=True)

# Sample rows per governorate for demo
sample_entries = {
    "الكوت": [
        ("1-Kut", "الكوت", "LDP", "Paracetamol", 500, "2026-08-15"),
        ("2-Kut", "الكوت", "MEDREICH", "Amoxicillin", 400, "2026-08-16"),
        ("مذخر الكوت مباشر", "الكوت", "LDP", "Azithromycin", 800, "2026-08-17"),
        ("3-Kut", "الكوت", "MEDREICH", "Omeprazole", 350, "2026-08-18"),
        ("4-Kut", "الكوت", "LDP", "Amlodipine", 600, "2026-08-19"),
    ],
    "العمارة": [
        ("1-Amara", "العمارة", "LDP", "Paracetamol", 400, "2026-08-15"),
        ("2-Amara", "العمارة", "MEDREICH", "Amoxicillin", 300, "2026-08-16"),
        ("مذخر العمارة مباشر", "العمارة", "LDP", "Ibuprofen", 700, "2026-08-17"),
        ("3-Amara", "العمارة", "MEDREICH", "Atorvastatin", 250, "2026-08-18"),
    ],
    "البصرة": [
        ("1-Basra", "البصرة", "LDP", "Azithromycin", 1200, "2026-08-15"),
        ("2-Basra", "البصرة", "MEDREICH", "Amoxicillin", 800, "2026-08-16"),
        ("مذخر البصرة مباشر", "البصرة", "LDP", "Paracetamol", 2000, "2026-08-17"),
        ("3-Basra", "البصرة", "MEDREICH", "Atorvastatin", 600, "2026-08-18"),
        ("4-Basra", "البصرة", "LDP", "Ciprofloxacin", 900, "2026-08-19"),
    ],
    "الناصرية": [
        ("1-Nasiriya", "الناصرية", "LDP", "Paracetamol", 600, "2026-08-15"),
        ("2-Nasiriya", "الناصرية", "MEDREICH", "Amoxicillin", 500, "2026-08-16"),
        ("مذخر الناصرية مباشر", "الناصرية", "LDP", "Diclofenac", 1000, "2026-08-17"),
        ("3-Nasiriya", "الناصرية", "MEDREICH", "Pantoprazole", 400, "2026-08-18"),
    ],
}

for sheet_title, gov_name, delegate_list_str, default_target in gov_configs:
    ws = wb.create_sheet(title=sheet_title)
    ws.views.sheetView[0].rightToLeft = True

    # 1. Summary KPI Header Card (Rows 1 & 2)
    ws.merge_cells("A1:C1")
    ws.merge_cells("A2:C2")
    ws["A1"] = f"سجل مبيعات ومندوبي محافظة {gov_name}"
    ws["A1"].font = title_font
    ws["A1"].alignment = right_align

    ws["A2"] = f"الهدف المالي الشهري: {default_target:,} د.ع"
    ws["A2"].font = italic_font
    ws["A2"].alignment = right_align

    # Mini KPI block on top right
    ws["D1"] = "إجمالي المبيعات"
    ws["D1"].font = kpi_title_font
    ws["D1"].fill = teal_light_fill
    ws["D1"].alignment = center_align
    ws["D1"].border = all_border

    ws["D2"] = "=SUM(G5:G100)"
    ws["D2"].font = kpi_val_font
    ws["D2"].number_format = '#,##0 "د.ع"'
    ws["D2"].alignment = center_align
    ws["D2"].border = all_border

    ws["E1"] = "إجمالي القطع"
    ws["E1"].font = kpi_title_font
    ws["E1"].fill = teal_light_fill
    ws["E1"].alignment = center_align
    ws["E1"].border = all_border

    ws["E2"] = "=SUM(E5:E100)"
    ws["E2"].font = kpi_val_font
    ws["E2"].number_format = '#,##0 "قطعة"'
    ws["E2"].alignment = center_align
    ws["E2"].border = all_border

    ws["F1"] = "نسبة تحقيق التارجت"
    ws["F1"].font = kpi_title_font
    ws["F1"].fill = green_soft_fill
    ws["F1"].alignment = center_align
    ws["F1"].border = all_border

    ws["F2"] = f"=IF(I5>0, D2/I5, D2/{default_target})"
    ws["F2"].font = kpi_val_font
    ws["F2"].number_format = '0.0%'
    ws["F2"].alignment = center_align
    ws["F2"].border = all_border

    # Blank row 3
    ws.row_dimensions[3].height = 10

    # 2. Main Data Table Headers (Row 4)
    headers = [
        "المندوب",
        "المحافظة",
        "الشركة",
        "المادة",
        "العدد",
        "سعر المادة",
        "اجمالي المبلغ",
        "التاريخ",
        "تاركت المحافظة",
        "نسبه الانجاز",
    ]
    ws.row_dimensions[4].height = 28
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(4, col_idx, h)
        cell.font = header_font
        cell.fill = slate_header_fill
        cell.alignment = center_align
        cell.border = all_border

    # Data Validations for this sheet
    del_val = DataValidation(type="list", formula1=f'"{delegate_list_str}"', allow_blank=True)
    ws.add_data_validation(del_val)
    del_val.add("A5:A100")

    ws.add_data_validation(company_val)
    company_val.add("C5:C100")

    ws.add_data_validation(item_val)
    item_val.add("D5:D100")

    # Sample rows
    samples = sample_entries.get(gov_name, [])
    current_row = 5
    for s in samples:
        del_name, gov_val, comp, item_name, qty, dt = s
        
        ws.cell(current_row, 1, del_name) # المندوب
        ws.cell(current_row, 2, gov_val)  # المحافظة
        ws.cell(current_row, 3, comp)     # الشركة
        ws.cell(current_row, 4, item_name)# المادة
        ws.cell(current_row, 5, qty)      # العدد
        
        # سعر المادة Formula
        ws.cell(current_row, 6, f"=IFERROR(VLOOKUP(D{current_row}, 'المواد والأسعار'!$A$2:$B$20, 2, FALSE), 0)")
        
        # اجمالي المبلغ Formula
        ws.cell(current_row, 7, f"=E{current_row}*F{current_row}")
        
        # التاريخ
        ws.cell(current_row, 8, dt)
        
        # تاركت المحافظة Formula (Matches target from 'تارجت المحافظات' or fallback)
        ws.cell(current_row, 9, f"=IFERROR(SUMIFS('تارجت المحافظات'!$D$2:$D$50, 'تارجت المحافظات'!$C$2:$C$50, B{current_row}, 'تارجت المحافظات'!$B$2:$B$50, MONTH(H{current_row})), {default_target})")
        
        # نسبه الانجاز Formula
        ws.cell(current_row, 10, f"=IF(I{current_row}>0, G{current_row}/I{current_row}, 0)")

        # Styling row
        fill = zebra_fill if current_row % 2 == 0 else PatternFill(fill_type=None)
        for c in range(1, 11):
            cell = ws.cell(current_row, c)
            cell.font = regular_font
            cell.border = all_border
            if fill.fill_type:
                cell.fill = fill
            if c in [1, 2, 3, 4]:
                cell.alignment = right_align
            elif c in [5, 8]:
                cell.alignment = center_align
            elif c in [6, 7, 9]:
                cell.alignment = left_align
                cell.number_format = '#,##0 "د.ع"'
                if c == 7:
                    cell.font = bold_font
            elif c == 10:
                cell.alignment = center_align
                cell.number_format = '0.0%'
                cell.font = bold_font
                
        ws.row_dimensions[current_row].height = 24
        current_row += 1

    # Add 15 prepared empty rows with formulas ready
    for r_empty in range(current_row, current_row + 15):
        ws.cell(r_empty, 2, gov_name) # Auto default governorate
        ws.cell(r_empty, 6, f"=IFERROR(VLOOKUP(D{r_empty}, 'المواد والأسعار'!$A$2:$B$20, 2, FALSE), 0)")
        ws.cell(r_empty, 7, f"=IF(E{r_empty}>0, E{r_empty}*F{r_empty}, 0)")
        ws.cell(r_empty, 9, f"=IFERROR(SUMIFS('تارجت المحافظات'!$D$2:$D$50, 'تارجت المحافظات'!$C$2:$C$50, B{r_empty}, 'تارجت المحافظات'!$B$2:$B$50, MONTH(H{r_empty})), {default_target})")
        ws.cell(r_empty, 10, f"=IF(I{r_empty}>0, G{r_empty}/I{r_empty}, 0)")
        
        fill = zebra_fill if r_empty % 2 == 0 else PatternFill(fill_type=None)
        for c in range(1, 11):
            cell = ws.cell(r_empty, c)
            cell.font = regular_font
            cell.border = all_border
            if fill.fill_type:
                cell.fill = fill
            if c in [6, 7, 9]:
                cell.number_format = '#,##0 "د.ع"'
            elif c == 10:
                cell.number_format = '0.0%'
        ws.row_dimensions[r_empty].height = 22

    # Totals Summary Row at the end
    tot_row = current_row + 15
    ws.cell(tot_row, 1, "الإجمالي العام للمحافظة").font = bold_font
    ws.cell(tot_row, 1).alignment = center_align
    ws.merge_cells(f"A{tot_row}:D{tot_row}")
    
    ws.cell(tot_row, 5, f"=SUM(E5:E{tot_row-1})").font = bold_font # إجمالي العدد
    ws.cell(tot_row, 5).alignment = center_align
    ws.cell(tot_row, 5).number_format = '#,##0'
    
    ws.cell(tot_row, 7, f"=SUM(G5:G{tot_row-1})").font = bold_font # إجمالي المبيعات
    ws.cell(tot_row, 7).number_format = '#,##0 "د.ع"'
    ws.cell(tot_row, 7).fill = highlight_yellow
    
    ws.cell(tot_row, 9, f"=MAX(I5:I{tot_row-1})").font = bold_font # التارغت
    ws.cell(tot_row, 9).number_format = '#,##0 "د.ع"'
    
    ws.cell(tot_row, 10, f"=IF(I{tot_row}>0, G{tot_row}/I{tot_row}, 0)").font = bold_font # نسبة الإنجاز الكلية
    ws.cell(tot_row, 10).number_format = '0.0%'
    ws.cell(tot_row, 10).fill = green_soft_fill
    
    for c in range(1, 11):
        ws.cell(tot_row, c).border = total_border
    ws.row_dimensions[tot_row].height = 26

# ==============================================================================
# SHEET 8: لوحة التحكم والملخص العام (Executive Dashboard)
# ==============================================================================
ws_dash = wb.create_sheet(title="لوحة التحكم العامة", index=0)
ws_dash.views.sheetView[0].rightToLeft = True

ws_dash.merge_cells("A1:G1")
ws_dash["A1"] = "نظام نبع الغدير العلمي — لوحة متابعة المندوبين والمحافظات"
ws_dash["A1"].font = Font(name=font_family, size=15, bold=True, color="FFFFFF")
ws_dash["A1"].fill = teal_dark_fill
ws_dash["A1"].alignment = center_align
ws_dash.row_dimensions[1].height = 36

# Summary Table Header (Row 3)
dash_headers = [
    "المحافظة",
    "عدد المندوبين",
    "إجمالي القطع المباعة",
    "إجمالي المبيعات المحققة",
    "الهدف الشهري (التارغت)",
    "نسبة الإنجاز %",
    "الحالة",
]
ws_dash.row_dimensions[3].height = 26
for c_idx, dh in enumerate(dash_headers, start=1):
    c = ws_dash.cell(3, c_idx, dh)
    c.font = header_font
    c.fill = slate_header_fill
    c.alignment = center_align
    c.border = all_border

gov_dash_rows = [
    ("الكوت", "سجلات الكوت", 8, 25000000),
    ("العمارة", "سجلات العمارة", 8, 20000000),
    ("البصرة", "سجلات البصرة", 8, 35000000),
    ("الناصرية", "سجلات الناصرية", 8, 22000000),
]

for idx, (g_name, s_name, del_cnt, tgt) in enumerate(gov_dash_rows, start=4):
    ws_dash.cell(idx, 1, g_name).alignment = center_align
    ws_dash.cell(idx, 1).font = bold_font
    
    ws_dash.cell(idx, 2, del_cnt).alignment = center_align
    
    # القطع المباعة Formula
    ws_dash.cell(idx, 3, f"='{s_name}'!E2").number_format = '#,##0 "قطعة"'
    ws_dash.cell(idx, 3).alignment = center_align
    ws_dash.cell(idx, 3).font = bold_font
    
    # المبيعات Formula
    ws_dash.cell(idx, 4, f"='{s_name}'!D2").number_format = '#,##0 "د.ع"'
    ws_dash.cell(idx, 4).alignment = left_align
    ws_dash.cell(idx, 4).font = bold_font
    
    # الهدف Formula
    ws_dash.cell(idx, 5, tgt).number_format = '#,##0 "د.ع"'
    ws_dash.cell(idx, 5).alignment = left_align
    
    # نسبة الإنجاز Formula
    ws_dash.cell(idx, 6, f"=IF(E{idx}>0, D{idx}/E{idx}, 0)").number_format = '0.0%'
    ws_dash.cell(idx, 6).alignment = center_align
    ws_dash.cell(idx, 6).font = bold_font
    ws_dash.cell(idx, 6).fill = teal_light_fill
    
    # الحالة Formula
    ws_dash.cell(idx, 7, f'=IF(F{idx}>=1, "محقق بنجاح 🎯", IF(F{idx}>=0.5, "قيد الإنجاز ⚡", "متأخر عن الخطة ⚠️"))').alignment = center_align
    ws_dash.cell(idx, 7).font = bold_font

    fill = zebra_fill if idx % 2 == 0 else PatternFill(fill_type=None)
    for c in range(1, 8):
        cell = ws_dash.cell(idx, c)
        cell.border = all_border
        if fill.fill_type and c != 6:
            cell.fill = fill
    ws_dash.row_dimensions[idx].height = 24

# Total Row in Dashboard
t_row = 8
ws_dash.cell(t_row, 1, "المجموع الكلي لجميع المحافظات").font = bold_font
ws_dash.cell(t_row, 1).alignment = center_align
ws_dash.merge_cells(f"A{t_row}:B{t_row}")

ws_dash.cell(t_row, 3, f"=SUM(C4:C7)").number_format = '#,##0 "قطعة"'
ws_dash.cell(t_row, 3).font = bold_font
ws_dash.cell(t_row, 3).alignment = center_align

ws_dash.cell(t_row, 4, f"=SUM(D4:D7)").number_format = '#,##0 "د.ع"'
ws_dash.cell(t_row, 4).font = bold_font
ws_dash.cell(t_row, 4).alignment = left_align
ws_dash.cell(t_row, 4).fill = highlight_yellow

ws_dash.cell(t_row, 5, f"=SUM(E4:E7)").number_format = '#,##0 "د.ع"'
ws_dash.cell(t_row, 5).font = bold_font
ws_dash.cell(t_row, 5).alignment = left_align

ws_dash.cell(t_row, 6, f"=IF(E{t_row}>0, D{t_row}/E{t_row}, 0)").number_format = '0.0%'
ws_dash.cell(t_row, 6).font = bold_font
ws_dash.cell(t_row, 6).alignment = center_align
ws_dash.cell(t_row, 6).fill = green_soft_fill

ws_dash.cell(t_row, 7, f'=IF(F{t_row}>=1, "إجمالي ممتاز", "متابعة مستمرة")').font = bold_font
ws_dash.cell(t_row, 7).alignment = center_align

for c in range(1, 8):
    ws_dash.cell(t_row, c).border = total_border
ws_dash.row_dimensions[t_row].height = 28

# Auto-adjust column widths for all sheets
for ws_curr in wb.worksheets:
    for col in ws_curr.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if val_str.startswith('='):
                val_str = "1234567890" # approximate formula length
            max_len = max(max_len, len(val_str))
        ws_curr.column_dimensions[col_letter].width = max(max_len + 5, 14)

output_file = "نظام_نبع_الغدير_المندوبين_والمحافظات.xlsx"
wb.save(output_file)
print(f"File successfully created: {output_file}")
